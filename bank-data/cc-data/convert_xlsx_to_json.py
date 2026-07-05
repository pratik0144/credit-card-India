#!/usr/bin/env python3
"""
Convert axisfinaldone18jun2026.xlsx → Master-data-banks.json

• Reads every cell from the Excel file.
• Strips wrapping single-quotes from string values.
• Normalises None / empty cells to null in the JSON output.
• Writes a clean, pretty-printed JSON array of objects
  (one object per row, keyed by the header row).
"""

import json
import openpyxl
from datetime import datetime, date, time
from pathlib import Path

XLSX_PATH = Path(__file__).parent / "axisfinaldone18jun2026.xlsx"
JSON_PATH = Path(__file__).parent / "Master-data-banks.json"


def clean_value(val):
    """Return a JSON-safe, cleaned version of a cell value."""
    if val is None:
        return None

    # openpyxl may return datetime / date / time objects
    if isinstance(val, datetime):
        return val.isoformat()
    if isinstance(val, date):
        return val.isoformat()
    if isinstance(val, time):
        return val.isoformat()

    if isinstance(val, str):
        # Strip wrapping single quotes that exist in every cell
        stripped = val.strip()
        if stripped.startswith("'") and stripped.endswith("'"):
            stripped = stripped[1:-1]
        stripped = stripped.strip()
        return stripped if stripped else None

    # int / float pass through
    return val


def main():
    wb = openpyxl.load_workbook(str(XLSX_PATH), data_only=True)
    ws = wb.active  # Sheet1

    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row,
                             max_col=ws.max_column, values_only=True))

    if not rows:
        print("⚠️  No data found in the workbook.")
        return

    # First row = headers
    headers = [clean_value(h) for h in rows[0]]
    print(f"✅ Found {len(headers)} columns, {len(rows) - 1} data rows.")

    records = []
    for row in rows[1:]:
        record = {}
        for col_idx, header in enumerate(headers):
            val = row[col_idx] if col_idx < len(row) else None
            record[header] = clean_value(val)
        records.append(record)

    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(records, f, indent=2, ensure_ascii=False)

    print(f"✅ Written {len(records)} records → {JSON_PATH}")


if __name__ == "__main__":
    main()
